from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


DESKTOP = Path(r"C:\Users\13398\Desktop")
INPUT_XLSX = DESKTOP / "mbart_150_error_analysis_with_reasons.xlsx"
LABELS = ["Preserved", "Omitted", "Literalized", "Other"]
TYPE_ORDER = ["kinship", "solidarity", "affectionate", "hierarchy", "insult"]


def normalize_label(value):
    text = "" if pd.isna(value) else str(value).strip()
    mapping = {
        "preserved": "Preserved",
        "preserve": "Preserved",
        "保留": "Preserved",
        "omitted": "Omitted",
        "omit": "Omitted",
        "省略": "Omitted",
        "literalized": "Literalized",
        "literalised": "Literalized",
        "literal": "Literalized",
        "字面化": "Literalized",
        "直译": "Literalized",
        "other": "Other",
        "其他": "Other",
        "mistranslated": "Other",
        "wrong": "Other",
    }
    return mapping.get(text.lower(), text)


def build_summaries(rows):
    rows = rows.copy()
    rows["final_error_label"] = rows["error_label"].apply(normalize_label)

    summary = (
        rows.pivot_table(index="称谓类型", columns="final_error_label", values="ID", aggfunc="count", fill_value=0)
        .reindex(TYPE_ORDER)
        .fillna(0)
        .astype(int)
    )
    for label in LABELS:
        if label not in summary.columns:
            summary[label] = 0
    summary = summary[LABELS]
    summary.loc["TOTAL"] = summary.sum(axis=0)
    summary = summary.reset_index().rename(columns={"称谓类型": "Type"})

    term_summary = (
        rows.pivot_table(index=["称谓类型", "称谓"], columns="final_error_label", values="ID", aggfunc="count", fill_value=0)
        .reset_index()
    )
    for label in LABELS:
        if label not in term_summary.columns:
            term_summary[label] = 0
    term_summary["条数"] = term_summary[LABELS].sum(axis=1)
    term_summary = term_summary[["称谓类型", "称谓", "条数"] + LABELS]
    return summary, term_summary


def replace_sheet(wb, sheet_name, df):
    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        ws_old = wb[sheet_name]
        wb.remove(ws_old)
        ws = wb.create_sheet(sheet_name, idx)
    else:
        ws = wb.create_sheet(sheet_name)

    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in col), 80)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)


def main():
    rows = pd.read_excel(INPUT_XLSX, sheet_name="row_analysis")
    summary, term_summary = build_summaries(rows)

    wb = load_workbook(INPUT_XLSX)
    replace_sheet(wb, "summary_matrix", summary)
    replace_sheet(wb, "term_summary", term_summary)
    wb.save(INPUT_XLSX)

    print(INPUT_XLSX)
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
